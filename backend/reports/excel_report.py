import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
import os

# ── Load Data ──────────────────────────────────────────────────────────────────
def load_data():
    df = pd.read_csv('exports/employees.csv')
    return df

# ── Styles ─────────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")
RED_FILL    = PatternFill("solid", fgColor="FF6B6B")
GREEN_FILL  = PatternFill("solid", fgColor="6BCB77")
TITLE_FONT  = Font(bold=True, size=14, color="1F4E79")

thin = Side(style='thin', color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

def style_data_rows(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row + 1):
        fill = ALT_FILL if r % 2 == 0 else PatternFill()
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center')

# ── Sheet 1: Raw Data ──────────────────────────────────────────────────────────
def sheet_raw_data(wb, df):
    ws = wb.create_sheet("Raw Data")
    headers = list(df.columns)
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    for _, row in df.iterrows():
        ws.append(list(row))
    style_data_rows(ws, 2, len(df) + 1, len(headers))
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.freeze_panes = "A2"
    print("Sheet 1: Raw Data done")

# ── Sheet 2: Department Pivot ──────────────────────────────────────────────────
def sheet_department_pivot(wb, df):
    ws = wb.create_sheet("Dept Analysis")
    pivot = df.groupby('department').agg(
        Headcount=('employee_id', 'count'),
        Avg_Salary=('salary', 'mean'),
        Attrition_Count=('attrition', 'sum'),
        Avg_Performance=('performance_score', 'mean')
    ).reset_index()
    pivot['Attrition_Rate_%'] = (pivot['Attrition_Count'] / pivot['Headcount'] * 100).round(2)
    pivot['Avg_Salary'] = pivot['Avg_Salary'].round(2)
    pivot['Avg_Performance'] = pivot['Avg_Performance'].round(2)

    ws['A1'] = "Department Analysis"
    ws['A1'].font = TITLE_FONT
    headers = ['Department', 'Headcount', 'Avg Salary (AED)', 'Attrition Count', 'Avg Performance', 'Attrition Rate %']
    ws.append(headers)
    style_header_row(ws, 2, len(headers))
    for _, row in pivot.iterrows():
        ws.append(list(row))
    style_data_rows(ws, 3, len(pivot) + 2, len(headers))

    # Color high attrition red
    for r in range(3, len(pivot) + 3):
        cell = ws.cell(row=r, column=6)
        if cell.value and float(cell.value) > 30:
            cell.fill = RED_FILL
        elif cell.value and float(cell.value) < 15:
            cell.fill = GREEN_FILL

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    # Bar chart
    chart = BarChart()
    chart.title = "Headcount by Department"
    chart.y_axis.title = "Headcount"
    chart.x_axis.title = "Department"
    data = Reference(ws, min_col=2, min_row=2, max_row=len(pivot) + 2)
    cats = Reference(ws, min_col=1, min_row=3, max_row=len(pivot) + 2)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 20
    chart.height = 12
    ws.add_chart(chart, "H3")
    print("Sheet 2: Dept Analysis done")

# ── Sheet 3: Salary Analysis ───────────────────────────────────────────────────
def sheet_salary_analysis(wb, df):
    ws = wb.create_sheet("Salary Analysis")
    ws['A1'] = "Salary Analysis by Department"
    ws['A1'].font = TITLE_FONT

    salary = df.groupby('department')['salary'].agg(['min', 'max', 'mean', 'median']).reset_index()
    salary.columns = ['Department', 'Min Salary', 'Max Salary', 'Avg Salary', 'Median Salary']
    salary = salary.round(2)

    ws.append(list(salary.columns))
    style_header_row(ws, 2, len(salary.columns))
    for _, row in salary.iterrows():
        ws.append(list(row))
    style_data_rows(ws, 3, len(salary) + 2, len(salary.columns))
    for col in range(1, len(salary.columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    print("Sheet 3: Salary Analysis done")

# ── Sheet 4: Attrition Breakdown ───────────────────────────────────────────────
def sheet_attrition(wb, df):
    ws = wb.create_sheet("Attrition Breakdown")
    ws['A1'] = "Attrition Breakdown"
    ws['A1'].font = TITLE_FONT

    att = df.groupby('department').agg(
        Total=('employee_id', 'count'),
        Left=('attrition', 'sum')
    ).reset_index()
    att['Stayed'] = att['Total'] - att['Left']
    att['Attrition Rate %'] = (att['Left'] / att['Total'] * 100).round(2)

    ws.append(['Department', 'Total', 'Left', 'Stayed', 'Attrition Rate %'])
    style_header_row(ws, 2, 5)
    for _, row in att.iterrows():
        ws.append(list(row))
    style_data_rows(ws, 3, len(att) + 2, 5)
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 22
    print("Sheet 4: Attrition Breakdown done")

# ── Sheet 5: Nationality Distribution ─────────────────────────────────────────
def sheet_nationality(wb, df):
    ws = wb.create_sheet("Nationality Distribution")
    ws['A1'] = "Nationality Distribution"
    ws['A1'].font = TITLE_FONT

    nat = df['nationality'].value_counts().reset_index()
    nat.columns = ['Nationality', 'Count']

    ws.append(['Nationality', 'Count'])
    style_header_row(ws, 2, 2)
    for _, row in nat.iterrows():
        ws.append(list(row))
    style_data_rows(ws, 3, len(nat) + 2, 2)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15

    # Pie chart
    chart = PieChart()
    chart.title = "Nationality Distribution"
    data = Reference(ws, min_col=2, min_row=2, max_row=len(nat) + 2)
    cats = Reference(ws, min_col=1, min_row=3, max_row=len(nat) + 2)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 18
    chart.height = 12
    ws.add_chart(chart, "D3")
    print("Sheet 5: Nationality Distribution done")

# ── Sheet 6: Executive Dashboard ──────────────────────────────────────────────
def sheet_dashboard(wb, df):
    ws = wb.create_sheet("Executive Dashboard")
    ws.sheet_properties.tabColor = "1F4E79"

    ws['A1'] = "TalentIQ UAE — Executive HR Dashboard"
    ws['A1'].font = Font(bold=True, size=16, color="1F4E79")
    ws.merge_cells('A1:D1')

    total_emp     = len(df)
    avg_salary    = round(df['salary'].mean(), 2)
    attrition_rate = round(df['attrition'].mean() * 100, 2)
    avg_perf      = round(df['performance_score'].mean(), 2)
    top_dept      = df['department'].value_counts().idxmax()
    top_nat       = df['nationality'].value_counts().idxmax()

    kpis = [
        ("Total Employees", total_emp),
        ("Average Salary (AED)", avg_salary),
        ("Attrition Rate %", attrition_rate),
        ("Avg Performance Score", avg_perf),
        ("Largest Department", top_dept),
        ("Top Nationality", top_nat),
    ]

    ws['A3'] = "KPI"
    ws['B3'] = "Value"
    style_header_row(ws, 3, 2)

    for i, (kpi, val) in enumerate(kpis, start=4):
        ws.cell(row=i, column=1, value=kpi).font = Font(bold=True)
        ws.cell(row=i, column=2, value=val)
        ws.cell(row=i, column=1).border = BORDER
        ws.cell(row=i, column=2).border = BORDER

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 20
    print("Sheet 6: Executive Dashboard done")

# ── Main ───────────────────────────────────────────────────────────────────────
def generate_excel_report():
    df = load_data()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    sheet_raw_data(wb, df)
    sheet_department_pivot(wb, df)
    sheet_salary_analysis(wb, df)
    sheet_attrition(wb, df)
    sheet_nationality(wb, df)
    sheet_dashboard(wb, df)

    os.makedirs('exports/excel', exist_ok=True)
    path = 'exports/excel/TalentIQ_HR_Report.xlsx'
    wb.save(path)
    print(f"\nExcel report saved to {path}")

if __name__ == "__main__":
    generate_excel_report()